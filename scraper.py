"""
Scraper de precios de aceite de oliva - Supermercados Argentina
Uso: python scraper.py
"""

import io
import json
import os
import re
import sys
import time
from datetime import date, datetime
from pathlib import Path

# Forzar UTF-8 en la consola (Windows cp1252 no soporta algunos caracteres)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from tracker_paths import historial_path, precios_db_path

# ---------------------------------------------------------------------------
# Configuración
# ---------------------------------------------------------------------------

DIRECTORIO = Path(__file__).parent
ARCHIVO_HISTORIAL = historial_path()
ARCHIVO_EXCEL = DIRECTORIO / "aceites_oliva_tracker.xlsx"

PRECIO_MIN = 2_000
PRECIO_MAX = 150_000

HEADERS_HTTP = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json",
}

TERMINOS_BUSQUEDA = [
    "aceite oliva extra virgen",
    "aceite de oliva",
    "aceite oliva",
]

PALABRAS_EXCLUIR = [
    "mayonesa", "hummus", "vinagre", "aderezo", "salsa", "pesto",
    "aceituna", "girasol", "maiz", "maíz", "soja", "canola", "spray",
    "atun", "atún", "papa", "papas",
]

PALABRAS_INCLUIR = [
    "oliva",
]

_HEADLESS_FALSE_VALUES = {"0", "false", "no", "off"}


def _resolver_headless(auto_mode: bool) -> bool:
    valor = os.environ.get("ACEITE_TRACKER_HEADLESS")
    if valor is None:
        return auto_mode
    return valor.strip().lower() not in _HEADLESS_FALSE_VALUES

VERDE_OFERTA = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
VERDE_HEADER = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
AZUL_HEADER  = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
GRIS_HEADER  = PatternFill(start_color="9E9E9E", end_color="9E9E9E", fill_type="solid")

# ---------------------------------------------------------------------------
# Utilidades generales
# ---------------------------------------------------------------------------

def es_aceite_oliva(nombre: str) -> bool:
    n = nombre.lower()
    # Requiere "oliva" Y "aceite" en el nombre
    if "oliva" not in n or "aceite" not in n:
        return False
    if any(p in n for p in ("queso", "quesos", "quesito", "quesitos")):
        return False
    if any(p in n for p in PALABRAS_EXCLUIR):
        return False
    return True


def _buscar_ml_en_texto(texto: str) -> int | None:
    """Busca ml/gr/litros en un texto libre. Devuelve siempre ml equivalentes."""
    t = texto.lower()

    # Patrón inverso: unidad ANTES del número (ej: "gr.-500", "bot-gr.-500")
    m = re.search(r"(?:grm[s]?|gr[s]?|gramo[s]?)\s*[.\-]+\s*(\d+)", t)
    if m:
        return int(m.group(1))

    # ml / cc / cmq(cm³) / gr / g / grm
    m = re.search(r"(\d+[\.,]?\d*)\s*(grm[s]?|gramo[s]?|gr[s]?|g(?=\b)|cmq|ml|cc)\b", t)
    if m:
        return int(float(m.group(1).replace(",", ".")))

    # Litros / kg  — incluye "ltr", "lt", "l" (cuidando no capturar palabras)
    m = re.search(r"(\d+[\.,]?\d*)\s*(kg[s]?|kilogramo[s]?|litro[s]?|ltr[s]?|lt[s]?|l(?![a-z]))\b", t)
    if m:
        return int(float(m.group(1).replace(",", ".")) * 1000)

    # "X 50" → 500ml (50cl); "X N≥100" → N ml
    m = re.search(r"\bx\s*(\d+)\b(?!\s*(?:ml|cc|gr?|l\b|lt|litro|unid|pack|u\b))", t)
    if m:
        n = int(m.group(1))
        if n == 50:
            return 500
        elif 100 <= n <= 5000:
            return n

    return None


def extraer_ml(nombre: str, measure_unit: str = "", unit_multiplier: float = 0.0,
               textos_extra: list[str] | None = None) -> int | None:
    """Extrae ml/gr del nombre, campos VTEX y textos adicionales."""
    # 1. Desde measurementUnit/unitMultiplier de VTEX
    if measure_unit and unit_multiplier:
        um = measure_unit.lower().strip()
        if um in ("ml", "cc", "g", "gr", "grs", "grm", "grms", "gramo", "gramos"):
            try:
                return int(float(unit_multiplier))
            except ValueError:
                pass
        elif um in ("l", "lt", "lts", "litro", "litros", "kg", "kgs", "kilogramo", "kilogramos"):
            try:
                return int(float(unit_multiplier) * 1000)
            except ValueError:
                pass

    # 2. Desde el nombre del producto
    resultado = _buscar_ml_en_texto(nombre)
    if resultado:
        return resultado

    # 3. Desde textos adicionales (descripción, nombre de SKU, especificaciones)
    for texto in (textos_extra or []):
        if texto:
            resultado = _buscar_ml_en_texto(texto)
            if resultado:
                return resultado

    return None


def precio_por_litro(precio: float, ml: int | None) -> int | None:
    if ml and ml > 0:
        return round(precio / ml * 1000)
    return None


def precio_valido(precio: float) -> bool:
    return PRECIO_MIN <= precio <= PRECIO_MAX


def formatear_pesos(valor) -> str:
    if valor is None:
        return ""
    return f"${valor:,.0f}".replace(",", ".")


def _float_or_zero(valor) -> float:
    try:
        return float(valor or 0)
    except (TypeError, ValueError):
        return 0.0


def _nombres_clusters_vtex(raw_clusters) -> list[str]:
    if isinstance(raw_clusters, dict):
        return [str(v).strip() for v in raw_clusters.values() if str(v).strip()]
    if isinstance(raw_clusters, list):
        nombres = []
        for cluster in raw_clusters:
            if isinstance(cluster, dict):
                nombre = (
                    cluster.get("name")
                    or cluster.get("Name")
                    or cluster.get("label")
                    or cluster.get("Label")
                    or ""
                )
            else:
                nombre = str(cluster)
            nombre = str(nombre).strip()
            if nombre:
                nombres.append(nombre)
        return nombres
    return []


def _teasers_vtex(item: dict, offer: dict) -> list[str]:
    textos = []
    for key in (
        "teasers",
        "Teasers",
        "promotionTeasers",
        "PromotionTeasers",
        "promotions",
        "Promotions",
    ):
        values = offer.get(key)
        if values is None:
            values = item.get(key)
        if not isinstance(values, list):
            continue
        for entry in values:
            if isinstance(entry, dict):
                nombre = (
                    entry.get("name")
                    or entry.get("Name")
                    or entry.get("title")
                    or entry.get("Title")
                    or entry.get("text")
                    or entry.get("Text")
                    or ""
                )
            else:
                nombre = str(entry)
            nombre = str(nombre).strip()
            if nombre:
                textos.append(nombre)
    return textos


def _parece_promocion_vtex(item: dict, offer: dict) -> bool:
    patrones_descuento = (
        r"\b\d{1,2}%\b",
        r"\boff\b",
        r"\bdto\b",
        r"\bdescuento\b",
        r"\bpromo\b",
    )
    textos = _nombres_clusters_vtex(item.get("productClusters", [])) + _teasers_vtex(item, offer)
    for texto in textos:
        t = texto.lower()
        if re.search(r"multi|segunda|2da|2x|3x", t):
            continue
        if any(re.search(patron, t) for patron in patrones_descuento):
            return True
    return False


def _resolver_precios_vtex(item: dict, offer: dict) -> tuple[float, float | None, bool]:
    price = _float_or_zero(offer.get("Price", 0))
    list_price = _float_or_zero(offer.get("ListPrice", 0))
    price_without_discount = _float_or_zero(
        offer.get("PriceWithoutDiscount", 0) or offer.get("priceWithoutDiscount", 0)
    )
    spot_price = _float_or_zero(offer.get("spotPrice", 0))

    if spot_price > 0 and spot_price < price * 0.99 and precio_valido(spot_price):
        return round(spot_price, 2), round(price, 2), True

    base_candidates = [
        valor
        for valor in (price_without_discount, list_price)
        if valor > price * 1.01 and precio_valido(valor)
    ]
    if not base_candidates:
        return round(price, 2), None, False

    precio_base = max(base_candidates)
    return round(price, 2), round(precio_base, 2), True


# ---------------------------------------------------------------------------
# Scrapers VTEX
# ---------------------------------------------------------------------------

VTEX_SUPERS = {
    "Carrefour": "https://www.carrefour.com.ar",
    "Día":       "https://diaonline.supermercadosdia.com.ar",
}

# Cencosud: usan Intelligent Search API que devuelve más productos
CENCOSUD_SUPERS = {
    "Jumbo": "https://www.jumbo.com.ar",
    "Disco": "https://www.disco.com.ar",
    "Vea":   "https://www.vea.com.ar",
}

VTEX_CHANGO = [
    "https://www.masonline.com.ar",
    "https://www.changomas.com.ar",
]


def scrape_vtex(supermercado: str, base_url: str) -> list[dict]:
    productos = []
    vistos = set()

    for termino in TERMINOS_BUSQUEDA:
        url = (
            f"{base_url}/api/catalog_system/pub/products/search/"
            f"{requests.utils.quote(termino)}"
            f"?O=OrderByPriceDESC&_from=0&_to=47"
        )
        try:
            resp = requests.get(url, headers=HEADERS_HTTP, timeout=20)
            if resp.status_code not in (200, 206):
                continue
            data = resp.json()
            if not data:
                continue

            nuevos = 0
            for item in data:
                nombre = item.get("productName", "")
                if not es_aceite_oliva(nombre):
                    continue
                prod_id = str(item.get("productId", nombre))   # solo para dedup
                if prod_id in vistos:
                    continue
                vistos.add(prod_id)

                # URL real del producto (igual que scraper aceitunas)
                link = item.get("link") or ""
                if link and not link.startswith("http"):
                    link = base_url.rstrip("/") + link

                # Obtener precios desde el primer sku disponible
                skus = item.get("items", [])
                if not skus:
                    continue
                sku = skus[0]

                sellers = sku.get("sellers", [])
                if not sellers:
                    continue
                offer = sellers[0].get("commertialOffer", {})

                price, precio_sin, en_oferta = _resolver_precios_vtex(item, offer)
                disponible = int(offer.get("AvailableQuantity", 0) or 0)

                if price <= 0 or not precio_valido(price) or disponible <= 0:
                    continue

                measure_unit = sku.get("measurementUnit", "")
                unit_mult    = float(sku.get("unitMultiplier", 0) or 0)

                # Textos adicionales donde buscar tamaño
                textos_extra = [
                    sku.get("name", ""),
                    item.get("description", ""),
                    item.get("complementName", ""),
                    item.get("metaTagDescription", ""),
                ]
                # Especificaciones del producto (ej: "Contenido neto", "Peso")
                for spec_name in item.get("allSpecifications", []):
                    spec_vals = item.get(spec_name, [])
                    if isinstance(spec_vals, list):
                        textos_extra.extend(spec_vals)
                    elif isinstance(spec_vals, str):
                        textos_extra.append(spec_vals)

                ml = extraer_ml(nombre, measure_unit, unit_mult, textos_extra)

                productos.append({
                    "supermercado":   supermercado,
                    "nombre":         nombre,
                    "ml":             ml,
                    "precio":         round(price, 2),
                    "precio_sin_dto":  round(precio_sin, 2) if precio_sin else None,
                    "en_oferta":      en_oferta,
                    "producto_id":    link if link else prod_id,
                })
                nuevos += 1

            print(f"  [{supermercado}] '{termino}' → {nuevos} productos nuevos")

        except Exception as e:
            print(f"  [{supermercado}] Error con '{termino}': {e}")

    return productos


def scrape_changomas() -> list[dict]:
    """Usa la API de VTEX IO Intelligent Search de masonline.com.ar (sin Playwright)."""
    BASE = "https://www.masonline.com.ar/api/io/_v/api/intelligent-search/product_search/"
    productos = []
    vistos = set()
    desde = 0
    paso = 50

    while True:
        url = f"{BASE}?query=aceite+de+oliva&count={paso}&from={desde}&to={desde + paso - 1}"
        try:
            resp = requests.get(url, headers=HEADERS_HTTP, timeout=20)
            if resp.status_code not in (200, 206):
                break
            data = resp.json()
        except Exception as e:
            print(f"  [Chango Mas] Error en offset {desde}: {e}")
            break

        items_pagina = data.get("products", [])
        if not items_pagina:
            break

        nuevos = 0
        for item in items_pagina:
            nombre = item.get("productName", "")
            if not es_aceite_oliva(nombre):
                continue
            prod_id = item.get("productId", nombre)
            if prod_id in vistos:
                continue
            vistos.add(prod_id)

            skus = item.get("items", [])
            if not skus:
                continue
            offer = skus[0].get("sellers", [{}])[0].get("commertialOffer", {})
            price      = float(offer.get("Price", 0) or 0)
            list_price = float(offer.get("ListPrice", 0) or 0)

            if price <= 0 or not precio_valido(price):
                continue

            en_oferta  = list_price > price * 1.01 and precio_valido(list_price)
            precio_sin = round(list_price, 2) if en_oferta else None

            measure_unit = skus[0].get("measurementUnit", "")
            unit_mult    = float(skus[0].get("unitMultiplier", 0) or 0)
            textos_extra = [
                skus[0].get("name", ""),
                item.get("description", ""),
                item.get("complementName", ""),
            ]
            for spec_name in item.get("allSpecifications", []):
                spec_vals = item.get(spec_name, [])
                if isinstance(spec_vals, list):
                    textos_extra.extend(spec_vals)
                elif isinstance(spec_vals, str):
                    textos_extra.append(spec_vals)
            ml = extraer_ml(nombre, measure_unit, unit_mult, textos_extra)

            link = item.get("link") or ""
            productos.append({
                "supermercado":  "Chango Mas",
                "nombre":        nombre,
                "ml":            ml,
                "precio":        round(price, 2),
                "precio_sin_dto": precio_sin,
                "en_oferta":     en_oferta,
                "producto_id":   link if link.startswith("/") else str(prod_id),
            })
            nuevos += 1

        print(f"  [Chango Mas] offset {desde}: {nuevos} productos nuevos (total {len(productos)})")
        total = data.get("recordsFiltered", 0)
        desde += paso
        if desde >= total:
            break

    return productos


# ---------------------------------------------------------------------------
# Scraper Cencosud: Jumbo / Disco / Vea (API Intelligent Search)
# Usa el endpoint VTEX IO Intelligent Search con map=ft para obtener
# resultados reales de aceite de oliva (152 productos vs 48 del catalog).
# Scraper Playwright para Disco, Jumbo y Vea.
# Lee el precio renderizado en pantalla igual que Coto: captura precios
# de góndola y descuentos visibles (Fin de Semana, ListPrice, etc.)
# ---------------------------------------------------------------------------

def _parsear_precio_disco(texto: str) -> float | None:
    """Extrae precio en pesos de un string como '$20.350' o '$7.650'."""
    if not texto:
        return None
    corte = re.search(r"sin\s+impuesto", texto, re.IGNORECASE)
    if corte:
        texto = texto[:corte.start()]
    m = re.search(r'\$([\s\d.]+)', texto)
    if not m:
        return None
    try:
        v = float(m.group(1).replace(".", "").replace(",", ".").strip())
        if precio_valido(v):
            return v
    except ValueError:
        pass
    return None


# JS que corre dentro de la página Playwright para extraer productos con precios
def _leer_precio_publico_pdp_cencosud(page, product_url: str) -> tuple[float | None, float | None]:
    """
    Lee el precio visible en la PDP publica de Cencosud.
    Se usa para descartar ofertas fantasma cuando la grilla muestra un
    descuento que la pagina de detalle no confirma.
    """
    try:
        page.goto(product_url, timeout=30_000, wait_until="domcontentloaded")
        page.wait_for_timeout(4_000)
        texto = page.inner_text("body")
    except Exception:
        return None, None

    anchor = texto.find("SKU:")
    if anchor >= 0:
        texto = texto[anchor : anchor + 1_200]
    else:
        texto = texto[:1_200]

    precios: list[float] = []
    for linea in texto.splitlines():
        t = linea.strip()
        if not t or not t.startswith("$"):
            continue
        if re.search(r"lt\.|litro|impuesto", t, re.IGNORECASE):
            continue
        valor = _parsear_precio_disco(t)
        if valor and (not precios or abs(valor - precios[-1]) > 0.01):
            precios.append(valor)
        if len(precios) >= 2:
            break

    if not precios:
        return None, None

    precio_actual = precios[0]
    precio_original = None
    if len(precios) > 1 and precios[1] > precio_actual * 1.01:
        precio_original = precios[1]
    return precio_actual, precio_original


_JS_EXTRAER_CENCOSUD = r"""() => {
    // Contenedor: galleryItem (cada tarjeta de producto en la grilla VTEX)
    const cards = document.querySelectorAll('[class*="galleryItem"]');
    const result = [];
    for (const card of cards) {
        // Nombre del producto
        const nameEl = card.querySelector('[class*="nameContainer"], [class*="brandName"], [class*="productName"]');
        if (!nameEl) continue;
        const name = nameEl.textContent.trim();
        if (!name) continue;

        // Precios: recorrer TODOS los nodos hoja del card
        let currentPrice = null;
        let originalPrice = null;
        let discountPct = null;

        const allEls = card.querySelectorAll('*');
        for (const el of allEls) {
            if (el.children.length > 0) continue;
            const text = el.textContent.trim();

            // Badge de descuento tipo "-25%", "25%", "25% OFF"
            // Busca hasta 5 niveles de ancestro para capturar badges anidados
            if (!discountPct) {
                let clsCtx = '';
                let node = el;
                for (let i = 0; i < 5 && node; i++) {
                    clsCtx += ' ' + (node.getAttribute('class') || '');
                    node = node.parentElement;
                }
                const isBadge = /saving|discount|badge|promo|ofert|percent|sticker|label|tag/i.test(clsCtx);
                if (isBadge) {
                    const mPct = text.match(/^-?\s*(\d{1,2})\s*%/i);
                    if (mPct) {
                        const pct = parseInt(mPct[1]);
                        if (pct >= 5 && pct <= 80) discountPct = pct;
                    }
                }
            }

            // Precio argentino: empieza con $ seguido de dígitos y puntos
            if (!text.match(/^[$][0-9][0-9.]+$/)) continue;
            // Ignorar precio por litro ("Precio regular x lt.")
            const parent = el.parentElement;
            const parentText = parent ? parent.textContent : '';
            if (parentText.toLowerCase().includes('lt.') || parentText.toLowerCase().includes('litro')) continue;
            if (parentText.toLowerCase().includes('impuesto')) continue;

            // Detectar precio tachado: por CSS text-decoration O por nombre de clase
            const deco = window.getComputedStyle(el).textDecorationLine || '';
            let elClass = '';
            let node2 = el;
            for (let i = 0; i < 4 && node2; i++) {
                elClass += ' ' + (node2.getAttribute('class') || '');
                node2 = node2.parentElement;
            }
            const isStrike = deco.includes('line-through') ||
                /listPrice|list-price|priceWithout|without.?discount|sellingPrice.*strike|tachado|precio.?sin|regularPrice|compareAt/i.test(elClass);
            if (isStrike) {
                originalPrice = text;   // precio tachado = precio sin descuento
            } else {
                if (!currentPrice) currentPrice = text;   // primer precio sin tachado = precio a pagar
            }
        }

        const lines = (card.innerText || '')
            .split(/\n+/)
            .map(t => t.trim())
            .filter(Boolean);
        const priceLines = lines
            .filter(t => /^\$\s*\d/.test(t))
            .filter(t => !/(kg|lt\.|litro|impuesto)/i.test(t))
            .map(t => t.replace(/\s+/g, ''));
        if (!currentPrice && priceLines.length) {
            currentPrice = priceLines[0];
        }
        if (!originalPrice && priceLines.length > 1) {
            for (const candidate of priceLines.slice(1)) {
                if (candidate !== currentPrice) {
                    originalPrice = candidate;
                    break;
                }
            }
        }
        if (!discountPct) {
            for (const line of lines) {
                const mPct = line.match(/(?:^|\D)(\d{1,2})\s*%/i);
                if (!mPct) continue;
                const pct = parseInt(mPct[1]);
                if (pct >= 5 && pct <= 80) {
                    discountPct = pct;
                    break;
                }
            }
        }

        // Si hay badge de % pero no precio tachado: calcular el precio original
        if (currentPrice && discountPct && !originalPrice) {
            const num = parseFloat(currentPrice.slice(1).replace(/[.]/g, '').replace(',', '.'));
            const orig = Math.round(num / (1 - discountPct / 100));
            // Formatear como precio argentino (puntos de miles)
            originalPrice = '$' + orig.toLocaleString('es-AR');
        }

        // URL del producto (identificador único independiente del nombre)
        let productHref = null;
        const cardLink = card.querySelector('a[href]');
        if (cardLink) {
            const href = cardLink.getAttribute('href');
            if (href && href.startsWith('/')) productHref = href;
        }

        if (currentPrice) {
            result.push({ name, currentPrice, originalPrice, productHref });
        }
    }
    return result;
}"""


def scrape_cencosud(supermercado: str, base_url: str) -> list[dict]:
    """Scraper API para Disco, Jumbo y Vea via VTEX Intelligent Search."""
    BASE_IS = (
        f"{base_url}/api/io/_v/api/intelligent-search/product_search"
        "?query=aceite+oliva&count=50&map=ft&page={page}"
    )

    # Obtener cookies de sesión visitando la home — activa tablas de precios
    # (promotions tipo "Fin de Semana" sólo se reflejan en spotPrice con sesión)
    session = requests.Session()
    session.headers.update(HEADERS_HTTP)
    try:
        session.get(base_url, timeout=15, allow_redirects=True)
    except Exception:
        pass  # Si falla, igual intentamos sin cookies

    productos = []
    vistos = set()
    pagina = 1

    while True:
        url = BASE_IS.format(page=pagina)
        try:
            resp = session.get(url, timeout=20)
            if resp.status_code not in (200, 206):
                print(f"  [{supermercado}] HTTP {resp.status_code} en página {pagina}")
                break
            data = resp.json()
        except Exception as e:
            print(f"  [{supermercado}] Error en página {pagina}: {e}")
            break

        items_pagina = data.get("products", [])
        if not items_pagina:
            break

        nuevos = 0
        for item in items_pagina:
            nombre = item.get("productName", "")
            if not es_aceite_oliva(nombre):
                continue
            prod_id = item.get("productId", nombre)
            if prod_id in vistos:
                continue
            vistos.add(prod_id)

            skus = item.get("items", [])
            if not skus:
                continue
            sku = skus[0]
            offer = sku.get("sellers", [{}])[0].get("commertialOffer", {})

            price      = float(offer.get("Price", 0) or 0)
            list_price = float(offer.get("ListPrice", 0) or 0)
            spot_price = float(offer.get("spotPrice", 0) or 0)
            disponible = int(offer.get("AvailableQuantity", 0) or 0)

            if price <= 0 or not precio_valido(price) or disponible <= 0:
                continue

            # spotPrice < Price → precio de oferta real
            if spot_price > 0 and spot_price < price * 0.99 and precio_valido(spot_price):
                en_oferta  = True
                precio_sin = round(price, 2)
                price      = spot_price
            else:
                en_oferta  = list_price > price * 1.01 and precio_valido(list_price)
                precio_sin = round(list_price, 2) if en_oferta else None

            # productClusters: descuentos directos tipo "15% (20% con Cencopay)"
            # Se excluyen ofertas multi-unidad ("Hasta 2do al X%") y puntos ("Pesoscheck")
            if not en_oferta:
                clusters = item.get("productClusters", [])
                for c in clusters:
                    cname = c.get("name", "")
                    # Ignorar multi-unidad
                    if re.search(r'hasta\s+\d', cname, re.IGNORECASE):
                        continue
                    # Ignorar puntos/cashback
                    if re.search(r'pesoscheck|cashback|puntos', cname, re.IGNORECASE):
                        continue
                    # Capturar "15%" o "15% (20% con Cencopay)" → primer número
                    m = re.match(r'^(\d+)%', cname.strip())
                    if m:
                        pct = int(m.group(1))
                        if 1 <= pct <= 80:
                            en_oferta  = True
                            precio_sin = round(price, 2)
                            price      = round(price * (1 - pct / 100), 2)
                            break

            measure_unit = sku.get("measurementUnit", "")
            unit_mult    = float(sku.get("unitMultiplier", 0) or 0)
            textos_extra = [
                sku.get("name", ""),
                item.get("description", ""),
                item.get("complementName", ""),
                item.get("metaTagDescription", ""),
            ]
            for spec_name in item.get("allSpecifications", []):
                spec_vals = item.get(spec_name, [])
                if isinstance(spec_vals, list):
                    textos_extra.extend(spec_vals)
                elif isinstance(spec_vals, str):
                    textos_extra.append(spec_vals)

            ml = extraer_ml(nombre, measure_unit, unit_mult, textos_extra)

            productos.append({
                "supermercado":   supermercado,
                "nombre":         nombre,
                "ml":             ml,
                "precio":         round(price, 2),
                "precio_sin_dto": precio_sin,
                "en_oferta":      en_oferta,
                "producto_id":    str(prod_id),
            })
            nuevos += 1

        print(f"  [{supermercado}] página {pagina}: {nuevos} productos nuevos (total {len(productos)})")

        # Cortar solo cuando ya no hay más páginas según recordsFiltered
        # (NO cortar por nuevos==0: una página puede no tener aceite de oliva
        #  pero la siguiente sí puede tenerlo)
        total = data.get("recordsFiltered", 0)
        if total > 0 and pagina * 50 >= total:
            break
        if total == 0:
            break
        pagina += 1

    return productos


def scrape_cencosud_playwright(
    supermercado: str,
    base_url: str,
    headless: bool = False,
) -> list[dict]:
    """
    Scraper Playwright para Disco, Jumbo y Vea.
    Renderiza la página real y lee el precio visible + tachado (precio_sin_dto).
    Captura descuentos de Fin de Semana y cualquier otro descuento mostrado en pantalla.
    """
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print(f"  [{supermercado}] Playwright no instalado. Saltando.")
        return []

    BASE_URL = f"{base_url}/aceite%20de%20oliva?_q=ACEITE+DE+OLIVA&map=ft"
    productos = []
    vistos: set[str] = set()
    items_total: list[dict] = []

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=headless)
        page = browser.new_page()
        page_publica = None
        try:
            # ── Primar cookies VTEX ──────────────────────────────────────────
            # El cookie vtex_segment controla tabla de precios y promociones.
            # Sin él, Chromium arranca "sin canal" y no ve descuentos de fin
            # de semana (igual que Chrome recién abierto sin historial).
            # Solución: cargar la home primero para que VTEX setee el segmento,
            # igual que lo haría cualquier navegador con visitas previas.
            print(f"  [{supermercado}] Iniciando sesión VTEX (home)...")
            try:
                page.goto(base_url, timeout=60_000, wait_until="domcontentloaded")
                page.wait_for_timeout(3_000)   # dar tiempo a scripts de personalización
            except Exception as e_home:
                print(f"  [{supermercado}] Home tardó mucho, continuando igual...")

            pagina = 1
            while True:
                URL = BASE_URL if pagina == 1 else f"{BASE_URL}&page={pagina}"
                print(f"  [{supermercado}] Cargando página {pagina}: {URL}")
                page.goto(URL, timeout=30_000, wait_until="domcontentloaded")

                # Esperar que React renderice al menos una gallery card
                try:
                    page.wait_for_selector('[class*="galleryItem"]', timeout=20_000)
                except Exception:
                    print(f"  [{supermercado}] Sin products en pág {pagina}, terminando.")
                    break

                # Scroll para que todas las cards de esta página se rendericen
                for _ in range(20):
                    page.evaluate("window.scrollBy(0, 600)")
                    page.wait_for_timeout(300)
                page.evaluate("window.scrollTo(0, 0)")
                page.wait_for_timeout(800)

                n_cards = page.evaluate(
                    "document.querySelectorAll('[class*=\"galleryItem\"]').length"
                )
                print(f"  [{supermercado}] Pág {pagina}: {n_cards} cards")

                items_pag = page.evaluate(_JS_EXTRAER_CENCOSUD)
                items_total.extend(items_pag)

                if n_cards == 0 or not items_pag:
                    break
                pagina += 1
                if pagina > 10:   # máx 10 páginas por seguridad
                    break

        except Exception as e:
            print(f"  [{supermercado}] Error Playwright: {e}")
        finally:
            if page_publica is not None:
                page_publica.close()
            browser.close()

    for item in items_total:
        nombre = (item.get("name") or "").strip()
        if not nombre or not es_aceite_oliva(nombre):
            continue
        prod_key = item.get("productHref") or item.get("productId") or nombre
        if prod_key in vistos:
            continue
        vistos.add(prod_key)

        precio     = _parsear_precio_disco(item.get("currentPrice") or "")
        precio_orig = _parsear_precio_disco(item.get("originalPrice") or "")

        if not precio:
            continue

        en_oferta = bool(
            precio_orig
            and precio_orig > precio * 1.01
            and precio_valido(precio_orig)
        )

        if supermercado == "Vea" and en_oferta and item.get("productHref"):
            product_href = item.get("productHref") or ""
            product_url = f"{base_url}{product_href}" if product_href.startswith("/") else product_href
            try:
                from playwright.sync_api import sync_playwright
                with sync_playwright() as pw_publico:
                    browser_publico = pw_publico.chromium.launch(headless=True)
                    page_publica = browser_publico.new_page()
                    try:
                        pdp_precio, pdp_precio_orig = _leer_precio_publico_pdp_cencosud(
                            page_publica,
                            product_url,
                        )
                    finally:
                        browser_publico.close()
            except Exception:
                pdp_precio, pdp_precio_orig = None, None

            if (
                pdp_precio
                and pdp_precio > precio * 1.05
                and (not pdp_precio_orig or pdp_precio_orig <= pdp_precio * 1.01)
            ):
                print(
                    f"  [{supermercado}] Oferta descartada por PDP: "
                    f"{nombre[:70]} -> {formatear_pesos(pdp_precio)}"
                )
                precio = pdp_precio
                precio_orig = None
                en_oferta = False

        ml = extraer_ml(nombre)
        productos.append({
            "supermercado":   supermercado,
            "nombre":         nombre,
            "ml":             ml,
            "precio":         round(precio, 2),
            "precio_sin_dto": round(precio_orig, 2) if en_oferta else None,
            "en_oferta":      en_oferta,
            "producto_id":    prod_key,
        })

    print(f"  [{supermercado}] {len(productos)} productos de aceite de oliva")
    return productos


# Scraper Coto (Playwright)
# ---------------------------------------------------------------------------

def _parsear_precio_coto(texto: str) -> float | None:
    """Extrae el primer precio válido de un texto de Coto."""
    # Cortar antes de "sin impuesto"
    corte = re.search(r"sin\s+impuesto", texto, re.IGNORECASE)
    if corte:
        texto = texto[:corte.start()]
    matches = re.findall(r"\$[\s]*([\d.,]+)", texto)
    for m in matches:
        try:
            v = float(m.replace(".", "").replace(",", "."))
            if precio_valido(v):
                return v
        except ValueError:
            continue
    return None


_COTO_STEALTH_JS = """
Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
Object.defineProperty(navigator, 'platform', {get: () => 'Win32'});
Object.defineProperty(navigator, 'language', {get: () => 'es-AR'});
Object.defineProperty(navigator, 'languages', {get: () => ['es-AR', 'es', 'en-US', 'en']});
window.chrome = window.chrome || { runtime: {} };
"""


def _abrir_contexto_coto(pw, headless: bool):
    browser = pw.chromium.launch(headless=headless)
    context = browser.new_context(
        viewport={"width": 1400, "height": 1800},
        user_agent=HEADERS_HTTP["User-Agent"],
        locale="es-AR",
    )
    page = context.new_page()
    page.add_init_script(_COTO_STEALTH_JS)
    return browser, context, page


def _coto_bloqueado(page, html: str | None = None) -> bool:
    try:
        titulo = page.title().lower()
    except Exception:
        titulo = ""
    if "blocked" in titulo:
        return True
    contenido = (html or "").lower()
    return (
        "web page blocked" in contenido
        or "url you requested has been blocked" in contenido
    )


def scrape_coto(headless: bool = False) -> list[dict]:
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("  [Coto] Playwright no instalado. Saltando.")
        return []

    productos = []
    vistos = set()

    with sync_playwright() as pw:
        browser, context, page = _abrir_contexto_coto(pw, headless=headless)
        relanzado_visible = False

        pagina = 1
        while True:
            url = "https://www.cotodigital.com.ar/sitios/cdigi/productos/aceite-de-oliva"
            if pagina > 1:
                url += f"?page={pagina}"

            print(f"  [Coto] Pagina {pagina}...")
            try:
                page.goto(url, timeout=30000, wait_until="domcontentloaded")
                page.evaluate("window.scrollTo(0, 0)")
                time.sleep(1)
                # Scroll hasta que la altura no cambie (lazy-load completo)
                _coto_alt_ant = 0
                _coto_sin_cambio = 0
                for _ in range(60):                                   # máx ~30 seg
                    page.evaluate("window.scrollBy(0, 600)")
                    time.sleep(0.4)
                    _coto_alt_act = page.evaluate("document.body.scrollHeight")
                    if _coto_alt_act == _coto_alt_ant:
                        _coto_sin_cambio += 1
                        if _coto_sin_cambio >= 4:
                            break
                    else:
                        _coto_sin_cambio = 0
                    _coto_alt_ant = _coto_alt_act
                page.evaluate("window.scrollTo(0, 0)")
                time.sleep(1.5)
                html = page.content()
                if _coto_bloqueado(page, html):
                    if headless and not relanzado_visible:
                        print("  [Coto] Bloqueo detectado en headless. Reintentando en visible...")
                        context.close()
                        browser.close()
                        browser, context, page = _abrir_contexto_coto(pw, headless=False)
                        relanzado_visible = True
                        continue
                    print("  [Coto] Bloqueo de seguridad detectado. Abortando scrape.")
                    break
                n_items_coto = html.count("producto-card")
                if n_items_coto == 0 and headless and not relanzado_visible:
                    print(f"  [Coto] Pag {pagina} vacia en headless. Reintentando en visible...")
                    context.close()
                    browser.close()
                    browser, context, page = _abrir_contexto_coto(pw, headless=False)
                    relanzado_visible = True
                    continue
                print(f"  [Coto] Pag {pagina} cargada ({n_items_coto} cards en HTML)")
            except Exception as e:
                if headless and not relanzado_visible:
                    print(f"  [Coto] Headless fallo en pagina {pagina}: {e}")
                    print("  [Coto] Reintentando en visible...")
                    context.close()
                    browser.close()
                    browser, context, page = _abrir_contexto_coto(pw, headless=False)
                    relanzado_visible = True
                    continue
                print(f"  [Coto] Error en pagina {pagina}: {e}")
                break

            soup = BeautifulSoup(html, "html.parser")
            cards = soup.select(".producto-card")

            nuevos_pagina = 0
            for card in cards:
                # Nombre: .nombre-producto (el primero del card, no el del centro-precios)
                nombre_tag = card.select_one(".nombre-producto")
                if not nombre_tag:
                    continue
                nombre = nombre_tag.get_text(strip=True)
                if not es_aceite_oliva(nombre):
                    continue
                # href como ID único (distingue SKUs con mismo nombre de display)
                href_tag = card.select_one("a[href]")
                coto_id = href_tag["href"] if href_tag and href_tag.get("href") else nombre
                if coto_id in vistos:
                    continue
                vistos.add(coto_id)

                # Precio actual: .card-title
                precio_tag = card.select_one(".card-title")
                if not precio_tag:
                    continue
                precio_real = _parsear_precio_coto(precio_tag.get_text(" ", strip=True))
                if precio_real is None:
                    continue

                # Precio sin descuento: small con "Precio regular:"
                precio_sin = None
                en_oferta = False
                for small in card.select("small"):
                    texto_small = small.get_text(" ", strip=True)
                    if re.search(r"precio\s+regular", texto_small, re.IGNORECASE):
                        v = _parsear_precio_coto(texto_small)
                        if v and v > precio_real * 1.01:
                            precio_sin = v
                            en_oferta = True
                        break

                ml = extraer_ml(nombre)
                productos.append({
                    "supermercado":  "Coto",
                    "nombre":        nombre,
                    "ml":            ml,
                    "precio":        round(precio_real, 2),
                    "precio_sin_dto": round(precio_sin, 2) if precio_sin else None,
                    "en_oferta":     en_oferta,
                    "producto_id":   coto_id,
                })
                nuevos_pagina += 1

            print(f"  [Coto] Pagina {pagina} -> {nuevos_pagina} productos nuevos")
            if nuevos_pagina == 0:
                break
            pagina += 1
            time.sleep(1.5)

        context.close()
        browser.close()

    return productos


# Reemplazo robusto: reintenta la corrida completa de Coto y rechaza fotos parciales.
def _coto_umbral_minimo() -> int:
    try:
        import sqlite3

        if not DB_PATH.exists():
            return 70

        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        rows = cur.execute("""
            SELECT COUNT(*) AS n
            FROM precios
            WHERE supermercado = 'Coto' AND fecha < ?
            GROUP BY fecha
            HAVING n > 0
            ORDER BY fecha DESC
            LIMIT 7
        """, (str(date.today()),)).fetchall()
        conn.close()

        counts = sorted(int(r[0]) for r in rows if r and r[0])
        if not counts:
            return 70
        mediana = counts[len(counts) // 2]
        return max(70, int(mediana * 0.75))
    except Exception:
        return 70


def _scrape_coto_once_retry(headless: bool = False) -> tuple[list[dict], bool, str]:
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("  [Coto] Playwright no instalado. Saltando.")
        return [], False, "playwright no instalado"

    productos = []
    vistos = set()

    with sync_playwright() as pw:
        browser, context, page = _abrir_contexto_coto(pw, headless=headless)
        relanzado_visible = False
        corrida_completa = False
        motivo = "sin finalizar"

        try:
            pagina = 1
            while True:
                url = "https://www.cotodigital.com.ar/sitios/cdigi/productos/aceite-de-oliva"
                if pagina > 1:
                    url += f"?page={pagina}"

                print(f"  [Coto] Pagina {pagina}...")
                try:
                    page.goto(url, timeout=30000, wait_until="domcontentloaded")
                    page.evaluate("window.scrollTo(0, 0)")
                    time.sleep(1)
                    _coto_alt_ant = 0
                    _coto_sin_cambio = 0
                    for _ in range(60):
                        page.evaluate("window.scrollBy(0, 600)")
                        time.sleep(0.4)
                        _coto_alt_act = page.evaluate("document.body.scrollHeight")
                        if _coto_alt_act == _coto_alt_ant:
                            _coto_sin_cambio += 1
                            if _coto_sin_cambio >= 4:
                                break
                        else:
                            _coto_sin_cambio = 0
                        _coto_alt_ant = _coto_alt_act
                    page.evaluate("window.scrollTo(0, 0)")
                    time.sleep(1.5)
                    html = page.content()
                    if _coto_bloqueado(page, html):
                        if headless and not relanzado_visible:
                            print("  [Coto] Bloqueo detectado en headless. Reintentando en visible...")
                            context.close()
                            browser.close()
                            browser, context, page = _abrir_contexto_coto(pw, headless=False)
                            relanzado_visible = True
                            continue
                        motivo = "bloqueo de seguridad"
                        print("  [Coto] Bloqueo de seguridad detectado. Abortando scrape.")
                        break
                    n_items_coto = html.count("producto-card")
                    if n_items_coto == 0 and headless and not relanzado_visible:
                        print(f"  [Coto] Pag {pagina} vacia en headless. Reintentando en visible...")
                        context.close()
                        browser.close()
                        browser, context, page = _abrir_contexto_coto(pw, headless=False)
                        relanzado_visible = True
                        continue
                    print(f"  [Coto] Pag {pagina} cargada ({n_items_coto} cards en HTML)")
                except Exception as e:
                    if headless and not relanzado_visible:
                        print(f"  [Coto] Headless fallo en pagina {pagina}: {e}")
                        print("  [Coto] Reintentando en visible...")
                        context.close()
                        browser.close()
                        browser, context, page = _abrir_contexto_coto(pw, headless=False)
                        relanzado_visible = True
                        continue
                    motivo = f"error en pagina {pagina}: {e}"
                    print(f"  [Coto] Error en pagina {pagina}: {e}")
                    break

                soup = BeautifulSoup(html, "html.parser")
                cards = soup.select(".producto-card")

                nuevos_pagina = 0
                for card in cards:
                    nombre_tag = card.select_one(".nombre-producto")
                    if not nombre_tag:
                        continue
                    nombre = nombre_tag.get_text(strip=True)
                    if not es_aceite_oliva(nombre):
                        continue
                    href_tag = card.select_one("a[href]")
                    coto_id = href_tag["href"] if href_tag and href_tag.get("href") else nombre
                    if coto_id in vistos:
                        continue
                    vistos.add(coto_id)

                    precio_tag = card.select_one(".card-title")
                    if not precio_tag:
                        continue
                    precio_real = _parsear_precio_coto(precio_tag.get_text(" ", strip=True))
                    if precio_real is None:
                        continue

                    precio_sin = None
                    en_oferta = False
                    for small in card.select("small"):
                        texto_small = small.get_text(" ", strip=True)
                        if re.search(r"precio\s+regular", texto_small, re.IGNORECASE):
                            v = _parsear_precio_coto(texto_small)
                            if v and v > precio_real * 1.01:
                                precio_sin = v
                                en_oferta = True
                            break

                    ml = extraer_ml(nombre)
                    productos.append({
                        "supermercado":  "Coto",
                        "nombre":        nombre,
                        "ml":            ml,
                        "precio":        round(precio_real, 2),
                        "precio_sin_dto": round(precio_sin, 2) if precio_sin else None,
                        "en_oferta":     en_oferta,
                        "producto_id":   coto_id,
                    })
                    nuevos_pagina += 1

                print(f"  [Coto] Pagina {pagina} -> {nuevos_pagina} productos nuevos")
                if nuevos_pagina == 0:
                    corrida_completa = True
                    motivo = f"fin normal en pagina {pagina}"
                    break
                pagina += 1
                time.sleep(1.5)
        finally:
            context.close()
            browser.close()

    return productos, corrida_completa, motivo


def scrape_coto(headless: bool = False, max_attempts: int | None = None) -> list[dict]:
    max_attempts = max_attempts or (5 if headless else 3)
    umbral_minimo = _coto_umbral_minimo()
    mejor_cantidad = 0
    ultimo_motivo = "sin detalle"

    for intento in range(1, max_attempts + 1):
        if intento > 1:
            print(f"  [Coto] Reintentando corrida completa ({intento}/{max_attempts})...")
            time.sleep(min(4 * (intento - 1), 12))

        headless_intento = headless if intento == 1 else False
        productos, corrida_completa, motivo = _scrape_coto_once_retry(headless=headless_intento)
        ultimo_motivo = motivo
        mejor_cantidad = max(mejor_cantidad, len(productos))

        corrida_suficiente = len(productos) >= umbral_minimo
        pagina_interrumpida = None
        m_pagina = re.search(r"pagina\s+(\d+)", str(motivo), re.IGNORECASE)
        if m_pagina:
            pagina_interrumpida = int(m_pagina.group(1))
        corrida_parcial_usable = (
            not corrida_completa
            and corrida_suficiente
            and pagina_interrumpida is not None
            and pagina_interrumpida >= 5
        )
        if corrida_completa and corrida_suficiente:
            if intento > 1:
                print(f"  [Coto] Corrida valida en intento {intento} ({len(productos)} productos).")
            return productos
        if corrida_parcial_usable:
            print(
                f"  [Coto] Corrida parcial aceptada en intento {intento} "
                f"({len(productos)} productos hasta pagina {pagina_interrumpida})."
            )
            return productos

        fallas = []
        if not corrida_completa:
            fallas.append("corrida incompleta")
        if not corrida_suficiente:
            fallas.append(f"{len(productos)} productos (< umbral {umbral_minimo})")
        detalle = ", ".join(fallas) if fallas else motivo
        print(f"  [Coto] Intento {intento} descartado: {detalle}.")

    raise RuntimeError(
        f"Coto no pudo completarse tras {max_attempts} intentos. "
        f"Mejor intento: {mejor_cantidad} productos. Ultimo motivo: {ultimo_motivo}"
    )


# ---------------------------------------------------------------------------
# Scraper La Anónima (Playwright)
# ---------------------------------------------------------------------------

def _parse_numero_ar(s: str) -> float:
    """Parsea número argentino o US. '11040.00' → 11040.0, '11.040,00' → 11040.0"""
    s = s.strip()
    # Formato US con decimal: dígitos, un punto, exactamente 2 dígitos → float directo
    if re.match(r'^\d+\.\d{2}$', s):
        return float(s)
    # Formato argentino: 1.234,56 o 1.234
    return float(s.replace(".", "").replace(",", "."))


def _extraer_precio_anonima(texto: str) -> float | None:
    """Extrae el primer precio válido de un texto de La Anónima."""
    matches = re.findall(r"\$?\s*([\d.,]+)", texto)
    for m in matches:
        try:
            v = _parse_numero_ar(m)
            if precio_valido(v):
                return v
        except ValueError:
            continue
    return None

def scrape_anonima(headless: bool = False, _retry_count: int = 0) -> list[dict]:
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("  [La Anonima] Playwright no instalado. Saltando.")
        return []

    productos = []
    reintentar_visible = False

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=headless)
        page = browser.new_page()

        # Paso 1: Seleccionar sucursal via CP
        print("  [La Anonima] Configurando sucursal (CP 8400 - Bariloche)...")
        try:
            page.goto("https://www.laanonima.com.ar/", timeout=30000, wait_until="domcontentloaded")
            time.sleep(3)
            cp_input = page.query_selector("input[placeholder='Completa tu código postal']")
            if cp_input:
                cp_input.fill("8400")
                page.keyboard.press("Enter")
                time.sleep(4)
                try:
                    page.wait_for_selector("text=Bariloche", timeout=8000)
                    page.click("text=Bariloche")
                    time.sleep(3)
                    print("  [La Anonima] Sucursal Bariloche seleccionada")
                except Exception:
                    print("  [La Anonima] No aparecio modal de sucursales, continuando...")
            else:
                print("  [La Anonima] No se encontro input de CP, continuando...")
        except Exception as e:
            print(f"  [La Anonima] Error configurando sucursal: {e}")

        # Paso 2: Buscar aceite de oliva
        print("  [La Anonima] Buscando aceite de oliva...")
        try:
            page.goto(
                "https://www.laanonima.com.ar/buscar/aceite%20oliva",
                timeout=30000,
                wait_until="domcontentloaded",
            )
            time.sleep(3)
            try:
                page.wait_for_selector(".producto-item", timeout=15000)
            except Exception:
                time.sleep(6)
            # Scroll hasta que la altura no cambie (lazy-load completo)
            _altura_ant = 0
            _intentos_sin_cambio = 0
            for _ in range(60):                                  # máx ~30 seg
                page.evaluate("window.scrollBy(0, 600)")
                time.sleep(0.4)
                _altura_act = page.evaluate("document.body.scrollHeight")
                if _altura_act == _altura_ant:
                    _intentos_sin_cambio += 1
                    if _intentos_sin_cambio >= 4:               # 4 intentos sin cambio → terminó
                        break
                else:
                    _intentos_sin_cambio = 0
                _altura_ant = _altura_act
            # Scroll al top y esperar renderizado final
            page.evaluate("window.scrollTo(0, 0)")
            time.sleep(1.5)
            html = page.content()
            if "producto-item" not in html:
                time.sleep(6)
                html = page.content()
            n_items_visible = html.count("producto-item")
            print(f"  [La Anonima] Página cargada ({n_items_visible} items en HTML)")
        except Exception as e:
            print(f"  [La Anonima] Error buscando productos: {e}")
            browser.close()
            raise

        soup = BeautifulSoup(html, "html.parser")
        items = soup.select(".producto-item")
        extra_searches = [
            ("aceite", "https://www.laanonima.com.ar/buscar/aceite"),
            ("oliva", "https://www.laanonima.com.ar/buscar/oliva"),
            ("la toscana", "https://www.laanonima.com.ar/buscar/la%20toscana"),
            ("oliovita", "https://www.laanonima.com.ar/buscar/oliovita"),
            ("natura oliva", "https://www.laanonima.com.ar/buscar/natura%20oliva"),
        ]
        for search_label, search_url in extra_searches:
            try:
                page.goto(search_url, timeout=30000, wait_until="domcontentloaded")
                time.sleep(3)
                try:
                    page.wait_for_selector(".producto-item", timeout=12000)
                except Exception:
                    time.sleep(4)
                _altura_ant = 0
                _intentos_sin_cambio = 0
                for _ in range(60):
                    page.evaluate("window.scrollBy(0, 600)")
                    time.sleep(0.4)
                    _altura_act = page.evaluate("document.body.scrollHeight")
                    if _altura_act == _altura_ant:
                        _intentos_sin_cambio += 1
                        if _intentos_sin_cambio >= 4:
                            break
                    else:
                        _intentos_sin_cambio = 0
                    _altura_ant = _altura_act
                page.evaluate("window.scrollTo(0, 0)")
                time.sleep(1.5)
                html_extra = page.content()
                soup_extra = BeautifulSoup(html_extra, "html.parser")
                items_extra = soup_extra.select(".producto-item")
                items.extend(items_extra)
                print(f"  [La Anonima] '{search_label}' -> {len(items_extra)} items extra")
            except Exception as e:
                print(f"  [La Anonima] Error en busqueda extra '{search_label}': {e}")
        print(f"  [La Anonima] .producto-item acumulados: {len(items)}")

        vistos_an = set()
        for item in items:
            # ── Obtener nombre: h2.titulo (catálogo normal) o a[data-nombre] (buscador)
            nombre = ""
            titulo_tag = item.select_one("h2.titulo")
            if titulo_tag:
                nombre = titulo_tag.get_text(strip=True)
            else:
                a_tag = item.select_one("a[data-nombre]")
                if a_tag:
                    nombre = a_tag.get("data-nombre", "").strip()

            if not nombre or not es_aceite_oliva(nombre):
                continue
            # href como ID único
            a_link = item.select_one("a[href]")
            anon_id = a_link.get("href", nombre) if a_link else nombre
            if anon_id in vistos_an:
                continue
            vistos_an.add(anon_id)

            # ── Precio regular (el grande, ej: $ 13.800) ─────────────────
            precio_regular = None
            precio_div = item.select_one(".precio")
            if precio_div:
                primer_span = precio_div.select_one("span:not(.detalle-plus)")
                if primer_span:
                    precio_regular = _extraer_precio_anonima(primer_span.get_text(" ", strip=True))

            # ── Precio Plus / "Pagás c/u" (precio con oferta) ─────────────
            precio_plus = None
            plus_tag = item.select_one(".detalle-plus")
            if plus_tag:
                txt_plus = plus_tag.get_text(" ", strip=True)
                m_plus = re.search(r"pag[áa]s?\s*c/u\s*([\d.,]+)", txt_plus, re.IGNORECASE)
                if m_plus:
                    try:
                        v = _parse_numero_ar(m_plus.group(1))
                        if precio_valido(v):
                            precio_plus = v
                    except ValueError:
                        pass

            # ── Precio anterior / tachado ──────────────────────────────────
            precio_anterior = None
            ant_tag = item.select_one(".precio-anterior")
            if ant_tag:
                precio_anterior = _extraer_precio_anonima(ant_tag.get_text(" ", strip=True))

            # ── Resolver precio final ──────────────────────────────────────
            # Si hay Plus → es la oferta; regular es el precio sin oferta
            # Si hay precio_anterior → es el regular; precio_regular es el con oferta
            # Si solo hay precio_regular → sin oferta
            if precio_plus and precio_regular and precio_valido(precio_plus):
                precio_real = precio_plus
                precio_sin  = precio_anterior if precio_anterior and precio_anterior > precio_plus * 1.01 else precio_regular
                en_oferta   = True
            elif precio_regular and precio_valido(precio_regular):
                precio_real = precio_regular
                precio_sin  = precio_anterior if precio_anterior and precio_anterior > precio_regular * 1.01 else None
                en_oferta   = bool(precio_sin)
            else:
                continue

            ml = extraer_ml(nombre)
            productos.append({
                "supermercado":  "La Anonima",
                "nombre":        nombre,
                "ml":            ml,
                "precio":        round(precio_real, 2),
                "precio_sin_dto": round(precio_sin, 2) if precio_sin else None,
                "en_oferta":     en_oferta,
                "producto_id":   anon_id,
            })

        if not productos and headless and _retry_count < 1:
            print(f"  [La Anonima] Respuesta vacía. Reintentando scrape ({_retry_count + 1}/2)...")
            reintentar_visible = True

        print(f"  [La Anonima] {len(productos)} productos encontrados")
        browser.close()

    if reintentar_visible:
        time.sleep(3)
        return scrape_anonima(headless=False, _retry_count=_retry_count + 1)

    return productos


# ---------------------------------------------------------------------------
# Historial JSON + SQLite
# ---------------------------------------------------------------------------

DB_PATH = precios_db_path()

def _init_db(cur):
    cur.execute("""
        CREATE TABLE IF NOT EXISTS precios (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha          TEXT    NOT NULL,
            supermercado   TEXT    NOT NULL,
            nombre         TEXT    NOT NULL,
            ml             INTEGER,
            precio         REAL    NOT NULL,
            precio_sin_dto REAL,
            en_oferta      INTEGER NOT NULL DEFAULT 0,
            marca          TEXT,
            precio_litro   INTEGER,
            producto_id    TEXT,
            UNIQUE(fecha, supermercado, nombre)
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_fecha ON precios(fecha)")


def guardar_en_sqlite(productos: list[dict], fecha: str) -> None:
    import sqlite3
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    cur  = conn.cursor()
    _init_db(cur)
    # Solo borrar los supermercados que vienen en este lote — no pisar datos de otras cadenas
    _supermercados = list({p.get("supermercado", "") for p in productos if p.get("supermercado")})
    for _sup in _supermercados:
        cur.execute("DELETE FROM precios WHERE fecha = ? AND supermercado = ?", (fecha, _sup))
    cur.executemany("""
        INSERT OR IGNORE INTO precios
          (fecha, supermercado, nombre, ml, precio, precio_sin_dto,
           en_oferta, marca, precio_litro, producto_id)
        VALUES (:fecha, :supermercado, :nombre, :ml, :precio, :precio_sin_dto,
                :en_oferta, :marca, :precio_litro, :producto_id)
    """, [{"fecha": fecha, "supermercado": p.get("supermercado",""),
           "nombre": p.get("nombre",""), "ml": p.get("ml"),
           "precio": p.get("precio", 0), "precio_sin_dto": p.get("precio_sin_dto"),
           "en_oferta": 1 if p.get("en_oferta") else 0,
           "marca": p.get("marca"), "precio_litro": p.get("precio_litro"),
           "producto_id": p.get("producto_id")} for p in productos])
    conn.commit()
    conn.close()


def cargar_historial() -> dict:
    if ARCHIVO_HISTORIAL.exists():
        with open(ARCHIVO_HISTORIAL, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"semanas": []}


def guardar_historial(historial: dict) -> None:
    ARCHIVO_HISTORIAL.parent.mkdir(parents=True, exist_ok=True)
    with open(ARCHIVO_HISTORIAL, "w", encoding="utf-8") as f:
        json.dump(historial, f, ensure_ascii=False, indent=2)


def _conteos_recientes_cadena(tabla: str, cadena: str, limite: int = 7) -> list[int]:
    try:
        import sqlite3

        if not DB_PATH.exists():
            return []

        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        rows = cur.execute(f"""
            SELECT COUNT(*) AS n
            FROM {tabla}
            WHERE supermercado = ? AND fecha < ?
            GROUP BY fecha
            HAVING n > 0
            ORDER BY fecha DESC
            LIMIT ?
        """, (cadena, str(date.today()), limite)).fetchall()
        conn.close()
        return [int(r[0]) for r in rows if r and r[0]]
    except Exception:
        return []


def _umbral_reintento_cadena(tabla: str, cadena: str) -> int:
    counts = _conteos_recientes_cadena(tabla, cadena)
    if not counts:
        return 1
    counts = sorted(counts)
    mediana = counts[len(counts) // 2]
    if cadena == "La Anonima":
        # La Anonima cambio el buscador y hoy expone menos surtido por sucursal.
        # Mantenemos guardrail contra respuestas vacias, sin exigir el historico completo.
        return max(10, int(mediana * 0.30))
    return max(3, int(mediana * 0.65))


def _scrapear_cadena_con_recheck(
    cadena: str,
    tabla: str,
    etiqueta: str,
    scraper_fn,
    max_attempts: int = 3,
) -> list[dict]:
    recientes = _conteos_recientes_cadena(tabla, cadena)
    umbral = _umbral_reintento_cadena(tabla, cadena)
    mejor: list[dict] = []
    ultimo_error: Exception | None = None

    for intento in range(1, max_attempts + 1):
        if intento > 1:
            print(f"  [{cadena}] Recheck automatico ({intento}/{max_attempts})...")
            time.sleep(min(4 * (intento - 1), 12))

        try:
            prods = scraper_fn()
        except Exception as exc:
            ultimo_error = exc
            print(f"  [{cadena}] Error en intento {intento}: {exc}")
            continue
        n = len(prods)
        if n > len(mejor):
            mejor = prods

        ilogico = False
        motivo = ""
        if recientes:
            if n < umbral:
                ilogico = True
                motivo = f"{n} {etiqueta} (< umbral {umbral}, hist reciente {recientes[:3]})"
        elif n == 0:
            ilogico = True
            motivo = "0 resultados sin baseline previo"

        if not ilogico:
            return prods

        print(f"  [{cadena}] Resultado ilogico detectado: {motivo}")

    if ultimo_error and not mejor:
        raise RuntimeError(
            f"{cadena}: no se pudo validar la corrida tras {max_attempts} intentos "
            f"por errores consecutivos ({ultimo_error})."
        ) from ultimo_error

    raise RuntimeError(
        f"{cadena}: no se pudo validar la corrida tras {max_attempts} intentos. "
        f"Mejor intento: {len(mejor)} {etiqueta}."
    )


def agregar_corrida(historial: dict, productos: list[dict]) -> None:
    hoy = str(date.today())
    # Reemplazar si ya existe la fecha de hoy
    historial["semanas"] = [s for s in historial["semanas"] if s["fecha"] != hoy]
    historial["semanas"].append({"fecha": hoy, "productos": productos})


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

COLUMNAS_PRINCIPALES = [
    "Supermercado",     # 1
    "Marca",            # 2
    "Nombre",           # 3
    "Tamaño (ml)",      # 4
    "Precio c/Oferta",  # 5
    "Precio s/Oferta",  # 6
    "Precio/Litro",     # 7  (basado en precio real pagado)
    "¿En Oferta?",      # 8
]

FMT_PESOS = '"$"#,##0'

MESES_ES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
            "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

# Mapa alias (lowercase) → nombre canónico de marca
# Ordenado de mayor a menor longitud para evitar matches parciales
_MARCAS_ALIAS: dict[str, str] = {
    # Multi-palabra primero (más largo primero evita falsos parciales)
    "familia zuccardi":  "Familia Zuccardi",
    "filippo berio":     "Filippo Berio",
    "ciudad del lago":   "Ciudad Del Lago",
    "pietro coricelli":  "Pietro Coricelli",
    "cuisine & co":      "Cousine & Co",
    "cousine & co":      "Cousine & Co",
    "cuisine&co":        "Cousine & Co",
    "cousine&co":        "Cousine & Co",
    "pampa's gold":      "Pampa Gold",
    "pampa gold":        "Pampa Gold",
    "d.v. catena":       "DV Catena",
    "dv catena":         "DV Catena",
    "la toscana":        "La Toscana",
    "la española":       "La Española",
    "la espanola":       "La Española",
    "la riojana":        "La Riojana",
    "san huberto":       "San Huberto",
    "del monte":         "Del Monte",
    "de cecco":          "De Cecco",
    # Una palabra
    "zuccardi":          "Familia Zuccardi",
    "filippo":           "Filippo Berio",
    "cuisine":           "Cousine & Co",
    "cousine":           "Cousine & Co",
    "pietro":            "Pietro Coricelli",
    "costaflores":       "Costaflores",
    "yancanello":        "Yancanello",
    "valderrama":        "Valderrama",
    "terramater":        "Terramater",
    "carbonell":         "Carbonell",
    "colavita":          "Colavita",
    "fritolim":          "Fritolim",
    "rastrilla":         "Rastrilla",
    "cocinero":          "Cocinero",
    "oliovita":          "Oliovita",
    "kirkland":          "Kirkland",
    "olitalia":          "Olitalia",
    "cañuelas":          "Cañuelas",
    "yancanelo":         "Yancanello",
    "casalta":           "Casalta",
    "monini":            "Monini",
    "morixe":            "Morixe",
    "nucete":            "Nucete",
    "cortijo":           "Cortijo",
    "castell":           "Castell",
    "borges":            "Borges",
    "ybarra":            "Ybarra",
    "natura":            "Natura",
    "cecco":             "De Cecco",
    "magrì":             "Magrì",
    "magri":             "Magrì",
    "ekolo":             "Ekolo",
    "vigil":             "Vigil",
    "zuelo":             "Zuelo",
    "felix":             "Felix",
    "laur":              "Laur",
    "zucco":             "Zucco",
    "lopez":             "Lopez",
    "pisi":              "Pisi",
    "lira":              "Lira",
    "check":             "Check",
    "best":              "Best",
    "cook":              "Cook",
    "gallo":             "Gallo",
    "carm":              "Carm",
    "carrefour":         "Carrefour",
    "jumbo":             "Jumbo",
    "disco":             "Disco",
    "vea":               "Vea",
    "coto":              "Coto",
    "día":               "Día",
    "dia":               "Día",
}
_MARCAS_ALIAS_SORTED = sorted(_MARCAS_ALIAS.items(), key=lambda x: -len(x[0]))

# Palabras que NUNCA pueden ser una marca
_NO_MARCA = {
    # unidades de medida
    "ml", "cc", "lt", "lts", "ltr", "ltrs", "gr", "grm", "grms", "kg", "g", "l",
    # descriptores de producto
    "aceite", "oliva", "extra", "virgen", "virgen-extra", "extra-virgen",
    "extravirgen", "virgenextra", "organico", "orgánico",
    "organica", "clasico", "clásico", "clasica", "clásica", "suave", "intenso",
    "intensa", "medio", "blend", "picual", "arbequina", "koroneiki", "frantoio",
    "hojiblanca", "manzanilla", "temprana", "cosecha", "puro", "pura", "andino",
    "fuerte", "genovesa", "piscual", "tradicional", "premium", "seleccion",
    "selección", "especial", "natural", "sin", "tacc", "bío", "bio",
    # envases
    "botella", "lata", "envase", "bidon", "bidón", "tarro", "frasco", "pet",
    "vidrio", "aerosol", "doy", "pack", "brick", "rocío", "rocio",
    # conectores / genéricos
    "de", "en", "con", "sin", "el", "la", "los", "las", "un", "una", "por",
    "bot", "x", "y", "e", "o", "u", "a", "al",
}


def extraer_marca(nombre: str) -> str:
    n = nombre.lower()
    for alias, canonical in _MARCAS_ALIAS_SORTED:
        if alias in n:
            return canonical
    # Fallback: primera palabra no inválida, len > 1, sin dígitos
    for palabra in nombre.split():
        p = palabra.lower().strip(".,()-/&")
        if len(p) > 1 and p not in _NO_MARCA and not re.search(r'\d', p):
            return palabra.strip(".,()-/&")
    return "Otra"


def enriquecer_producto(p: dict) -> dict:
    precio = p["precio"]
    ml     = p.get("ml")
    # Respetar marca ya corregida; solo extraer si no existe
    marca = p.get("marca") or extraer_marca(p["nombre"])
    return {
        **p,
        "marca": marca,
        "precio_litro": precio_por_litro(precio, ml),
    }


def fila_producto(p: dict) -> list:
    return [
        p["supermercado"],
        p.get("marca", ""),
        p["nombre"],
        p.get("ml") if p.get("ml") else "S/D",
        int(round(p["precio"])),
        int(round(p["precio_sin_dto"])) if p.get("precio_sin_dto") else "",
        int(round(p["precio_litro"])) if p.get("precio_litro") else "",
        "Sí" if p.get("en_oferta") else "No",
    ]


def aplicar_header(ws, fila: int, columnas: list[str], fill: PatternFill) -> None:
    for col_idx, col_name in enumerate(columnas, 1):
        cell = ws.cell(row=fila, column=col_idx, value=col_name)
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def ajustar_columnas(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)


def borde_fino():
    lado = Side(style="thin", color="CCCCCC")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def aplicar_formato_precio(ws, cols_precio: list[int], fila_inicio: int) -> None:
    """Aplica formato $ a celdas numéricas de precio."""
    for row in ws.iter_rows(min_row=fila_inicio, max_row=ws.max_row):
        for col in cols_precio:
            cell = row[col - 1]
            if isinstance(cell.value, (int, float)):
                cell.number_format = FMT_PESOS


def aplicar_formato_condicional(ws, col_oferta: int, fila_inicio: int) -> None:
    """Verde para Sí, rojo para No en la columna ¿En Oferta?"""
    from openpyxl.formatting.rule import CellIsRule
    ROJO_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    col_letra = get_column_letter(col_oferta)
    rango = f"{col_letra}{fila_inicio}:{col_letra}{ws.max_row}"
    ws.conditional_formatting.add(
        rango,
        CellIsRule(operator="equal", formula=['"Sí"'], fill=VERDE_OFERTA)
    )
    ws.conditional_formatting.add(
        rango,
        CellIsRule(operator="equal", formula=['"No"'], fill=ROJO_FILL)
    )


def verificar_excel_cerrado() -> None:
    if not ARCHIVO_EXCEL.exists():
        return
    try:
        with open(ARCHIVO_EXCEL, "r+b"):
            pass
    except PermissionError:
        print(
            f"\nERROR: El archivo '{ARCHIVO_EXCEL.name}' está abierto en Excel.\n"
            "Cerralo antes de correr el script y volvé a intentar."
        )
        sys.exit(1)


def generar_excel(historial: dict) -> None:
    verificar_excel_cerrado()

    semanas = historial.get("semanas", [])
    if not semanas:
        print("Sin datos para generar Excel.")
        return

    wb = Workbook()

    # ── Hoja 1: Última Semana ──────────────────────────────────────────────
    ultima = semanas[-1]
    productos_ultima = [enriquecer_producto(p) for p in ultima["productos"]]

    ws1 = wb.active
    ws1.title = "Última Semana"
    ws1.append([f"Fecha: {ultima['fecha']}"])
    ws1["A1"].font = Font(bold=True, size=12)
    ws1.append([])
    aplicar_header(ws1, 3, COLUMNAS_PRINCIPALES, VERDE_HEADER)

    for p in productos_ultima:
        ws1.append(fila_producto(p))

    # cols: Super=1, Marca=2, Nombre=3, Tam=4, PrecioC=5, PrecioS=6, Litro=7, Oferta=8
    aplicar_formato_precio(ws1, cols_precio=[5, 6, 7], fila_inicio=4)
    aplicar_formato_condicional(ws1, col_oferta=8, fila_inicio=4)
    ajustar_columnas(ws1)

    # ── Hoja 2: Historial Completo ─────────────────────────────────────────
    # Cols: Fecha | Mes | Año | + COLUMNAS_PRINCIPALES
    ws2 = wb.create_sheet("Historial Completo")
    cols_hist = ["Fecha", "Mes", "Año"] + COLUMNAS_PRINCIPALES
    aplicar_header(ws2, 1, cols_hist, AZUL_HEADER)

    filas_hist = []
    for semana in semanas:
        fecha_str = semana["fecha"]
        try:
            dt = datetime.strptime(fecha_str, "%Y-%m-%d")
            mes_nombre = MESES_ES[dt.month - 1]
            anio = dt.year
        except ValueError:
            mes_nombre = ""
            anio = ""
        for p in semana["productos"]:
            pe = enriquecer_producto(p)
            filas_hist.append((fecha_str, p["supermercado"], [fecha_str, mes_nombre, anio] + fila_producto(pe)))

    filas_hist.sort(key=lambda x: (x[0], x[1]))
    for _, _, fila in filas_hist:
        ws2.append(fila)

    # cols: Fecha=1, Mes=2, Año=3, Super=4, Marca=5, Nombre=6, Tam=7,
    #        Precio c/O=8, Precio s/O=9, P/Litro=10, Oferta=11
    aplicar_formato_precio(ws2, cols_precio=[8, 9, 10], fila_inicio=2)
    aplicar_formato_condicional(ws2, col_oferta=11, fila_inicio=2)
    ajustar_columnas(ws2)

    # ── Hoja 3: Evolución sin Oferta ────────────────────────────────────────
    # Filas = (Fecha, Marca, Tamaño ml) unificados | Columnas = cadenas
    ws3 = wb.create_sheet("Evolución sin Oferta")

    cadenas_ord = sorted({p["supermercado"] for s in semanas for p in s["productos"]})
    header_evol = ["Fecha", "Marca", "Tamaño (ml)"] + cadenas_ord
    aplicar_header(ws3, 1, header_evol, GRIS_HEADER)

    filas_evol = []
    for semana in semanas:
        f = semana["fecha"]
        # Pivot: (marca, ml) → {cadena: precio_min}
        pivot: dict[tuple, dict] = {}
        for p in semana["productos"]:
            pe = enriquecer_producto(p)
            marca = pe.get("marca", "Otra")
            ml    = pe.get("ml") if pe.get("ml") else "S/D"
            cadena = p["supermercado"]
            precio_base = int(round(p.get("precio_sin_dto") or p["precio"]))
            clave = (marca, ml)
            if clave not in pivot:
                pivot[clave] = {}
            # Si hay varios del mismo brand+ml en la misma cadena, tomar el mínimo
            if cadena not in pivot[clave] or precio_base < pivot[clave][cadena]:
                pivot[clave][cadena] = precio_base

        for (marca, ml), precios_x_cadena in sorted(pivot.items(), key=lambda x: (str(x[0][0]), int(x[0][1]) if str(x[0][1]).isdigit() else 0)):
            fila = [f, marca, ml] + [precios_x_cadena.get(c, "") for c in cadenas_ord]
            filas_evol.append((f, marca, fila))

    # Ordenar por fecha luego marca
    filas_evol.sort(key=lambda x: (x[0], x[1]))
    for _, _, fila in filas_evol:
        ws3.append(fila)

    # Formato $ a columnas de cadenas
    col_cadena_start = 4
    n_cadenas = len(cadenas_ord)
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row,
                              min_col=col_cadena_start,
                              max_col=col_cadena_start + n_cadenas - 1):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = FMT_PESOS

    ajustar_columnas(ws3)

    wb.save(ARCHIVO_EXCEL)
    print(f"\nExcel guardado: {ARCHIVO_EXCEL}")


# ---------------------------------------------------------------------------
# Análisis de calidad (sin costo, Python puro)
# ---------------------------------------------------------------------------

ARCHIVO_ANALISIS = DIRECTORIO / "analisis_calidad.txt"

# Palabras que NUNCA son marcas válidas (para detección de errores)
_MARCAS_INVALIDAS = _NO_MARCA | {
    "aerosol", "rocio", "rocío", "vegetal", "virgen-extra",
    "golden", "otra", "medio", "ciudad",
}


def _segunda_pasada_marca(nombre: str) -> str:
    """Intento más agresivo de extraer marca para productos que quedaron como 'Otra'."""
    n = nombre.lower()
    # Buscar alias primero (ya cubre la mayoría)
    for alias, canonical in _MARCAS_ALIAS_SORTED:
        if alias in n:
            return canonical
    # Segunda pasada: acepta palabras de 3+ letras sin dígitos que no estén bloqueadas
    for palabra in nombre.split():
        p = palabra.lower().strip(".,()-/&°")
        if len(p) >= 3 and p not in _NO_MARCA and not re.search(r"\d", p):
            return palabra.strip(".,()-/&°").capitalize()
    return "Otra"


def analizar_calidad(todos_productos: list[dict]) -> list[dict]:
    """
    Analiza y AUTO-CORRIGE los datos scrapeados.
    Devuelve la lista con correcciones aplicadas y guarda un reporte.
    """
    print("\n[Análisis] Revisando y corrigiendo calidad de datos...")

    enriquecidos = [enriquecer_producto(p) for p in todos_productos]
    correcciones: list[str] = []

    # ── AUTO-CORRECCIÓN 1: re-asignar marca "Otra" con segunda pasada ────
    for p in enriquecidos:
        if p.get("marca") == "Otra":
            nueva = _segunda_pasada_marca(p["nombre"])
            if nueva != "Otra":
                correcciones.append(
                    f"Marca: '{p['nombre']}' → asignada '{nueva}' (era 'Otra')"
                )
                p["marca"] = nueva

    # ── AUTO-CORRECCIÓN 2: excluir productos con precio/litro imposible ──
    # (precio muy bajo por ml claramente erróneo; no se eliminan, se marca ml como None)
    for p in enriquecidos:
        pl = p.get("precio_litro")
        if pl and pl < 1_000:
            correcciones.append(
                f"ml inválido: '{p['nombre']}' ({p['supermercado']}) "
                f"ml={p.get('ml')} → precio/L=${int(pl):,} → ml reseteado"
            )
            p["ml"] = None
            p["precio_litro"] = None

    # ── DETECCIÓN (solo reporte, sin auto-fix) ────────────────────────────
    por_marca: dict[str, list] = {}
    sin_ml: list[dict] = []
    for p in enriquecidos:
        marca = p.get("marca", "Otra")
        por_marca.setdefault(marca, []).append(p)
        if not p.get("ml"):
            sin_ml.append(p)

    # Marcas que siguen siendo sospechosas tras correcciones
    marcas_sospechosas = [
        (marca, por_marca[marca][0]["nombre"])
        for marca in por_marca
        if (
            marca.lower() in _MARCAS_INVALIDAS
            or len(marca) <= 1
            or re.search(r"\d", marca)
            or marca == "Otra"
        )
    ]

    # Precios anómalos
    todos_precios = sorted(int(round(p["precio"])) for p in enriquecidos)
    mediana = todos_precios[len(todos_precios) // 2]
    precios_anomalos = [
        {"nombre": p["nombre"], "super": p["supermercado"],
         "precio": int(round(p["precio"])),
         "motivo": "muy bajo" if p["precio"] < mediana * 0.15 else "muy alto"}
        for p in enriquecidos
        if p["precio"] < mediana * 0.15 or p["precio"] > mediana * 5.0
    ]

    # Marcas duplicadas por capitalización distinta
    marcas_norm: dict[str, list] = {}
    for marca in por_marca:
        marcas_norm.setdefault(marca.lower().replace(" ", ""), []).append(marca)
    duplicados = {k: v for k, v in marcas_norm.items() if len(v) > 1}

    # Precio/litro extremo (post-corrección)
    litros_anomalos = [
        {"nombre": p["nombre"], "super": p["supermercado"],
         "ml": p.get("ml"), "precio_litro": int(round(p["precio_litro"]))}
        for p in enriquecidos
        if p.get("precio_litro") and p["precio_litro"] > 300_000
    ]

    # ── Armar reporte ─────────────────────────────────────────────────────
    fecha_str = date.today().strftime("%Y-%m-%d")
    lineas = [
        f"=== ANÁLISIS DE CALIDAD — {fecha_str} ===",
        f"Total: {len(enriquecidos)} productos | {len(por_marca)} marcas | {len(sin_ml)} sin ml",
        f"Precio mediana: ${mediana:,}",
        "",
        f"── CORRECCIONES APLICADAS ({len(correcciones)}) ──",
    ]
    lineas += ([f"  • {c}" for c in correcciones] if correcciones else ["  Ninguna."])

    lineas += ["", f"── MARCAS AÚN SOSPECHOSAS ({len(marcas_sospechosas)}) ──"]
    lineas += ([f"  • '{m}'  →  ej: {ej}" for m, ej in sorted(marcas_sospechosas)]
               if marcas_sospechosas else ["  Ninguna."])

    lineas += ["", f"── PRECIOS ANÓMALOS ({len(precios_anomalos)}) ──"]
    lineas += ([f"  • [{a['super']}] {a['nombre']} — ${a['precio']:,} ({a['motivo']})"
                for a in sorted(precios_anomalos, key=lambda x: x["precio"])]
               if precios_anomalos else ["  Ninguno."])

    lineas += ["", f"── MARCAS DUPLICADAS ({len(duplicados)}) ──"]
    lineas += ([f"  • {' / '.join(v)}" for v in duplicados.values()]
               if duplicados else ["  Ninguna."])

    lineas += ["", f"── PRODUCTOS SIN TAMAÑO ({len(sin_ml)}) ──"]
    lineas += [f"  • [{p['supermercado']}] {p['nombre']}" for p in sin_ml[:20]]
    if len(sin_ml) > 20:
        lineas.append(f"  ... y {len(sin_ml) - 20} más.")

    lineas += ["", f"── PRECIO/LITRO ANÓMALO ({len(litros_anomalos)}) ──"]
    lineas += ([f"  • [{a['super']}] {a['nombre']} — {a['ml']}ml → ${a['precio_litro']:,}/L"
                for a in litros_anomalos]
               if litros_anomalos else ["  Ninguno."])
    lineas.append("\n")

    modo = "a" if ARCHIVO_ANALISIS.exists() else "w"
    with open(ARCHIVO_ANALISIS, modo, encoding="utf-8") as f:
        f.write("\n".join(lineas))

    # Consola
    alertas = len(marcas_sospechosas) + len(precios_anomalos) + len(sin_ml)
    print(f"[Análisis] {len(correcciones)} correcciones aplicadas | {alertas} alertas restantes")
    if correcciones:
        print(f"  Corregidos: {len(correcciones)} productos")
    if marcas_sospechosas:
        print(f"  Marcas aún sospechosas: {', '.join(m for m, _ in marcas_sospechosas[:5])}")
    if sin_ml:
        print(f"  Sin tamaño: {len(sin_ml)} productos")
    print(f"  Reporte: {ARCHIVO_ANALISIS.name}")

    return enriquecidos


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    auto_mode = "--auto" in sys.argv or not sys.stdin.isatty()
    headless = _resolver_headless(auto_mode)

    print("=" * 60)
    print("Scraper de aceite de oliva - Supermercados Argentina")
    print(f"Fecha: {date.today()}")
    if auto_mode:
        print("Modo: automatico (sin verificacion interactiva)")
    print(f"Navegador: {'headless' if headless else 'visible'}")
    print("=" * 60)

    todos_los_productos: list[dict] = []

    # VTEX estándar (Carrefour, Día)
    for nombre, base_url in VTEX_SUPERS.items():
        print(f"\n[{nombre}]")
        prods = _scrapear_cadena_con_recheck(
            cadena=nombre,
            tabla="precios",
            etiqueta="productos",
            scraper_fn=lambda nombre=nombre, base_url=base_url: scrape_vtex(nombre, base_url),
        )
        print(f"  Total {nombre}: {len(prods)} productos")
        todos_los_productos.extend(prods)

    # Cencosud con Playwright (Jumbo, Disco, Vea) — lee precio renderizado incluyendo Fin de Semana
    for nombre, base_url in CENCOSUD_SUPERS.items():
        print(f"\n[{nombre}]")
        prods = _scrapear_cadena_con_recheck(
            cadena=nombre,
            tabla="precios",
            etiqueta="productos",
            scraper_fn=lambda nombre=nombre, base_url=base_url: scrape_cencosud_playwright(nombre, base_url, headless=headless),
        )
        print(f"  Total {nombre}: {len(prods)} productos")
        todos_los_productos.extend(prods)

    # Chango Más
    print("\n[Chango Más]")
    prods = _scrapear_cadena_con_recheck(
        cadena="Chango Mas",
        tabla="precios",
        etiqueta="productos",
        scraper_fn=scrape_changomas,
    )
    print(f"  Total Chango Más: {len(prods)} productos")
    todos_los_productos.extend(prods)

    # Coto
    print("\n[Coto]")
    prods = _scrapear_cadena_con_recheck(
        cadena="Coto",
        tabla="precios",
        etiqueta="productos",
        scraper_fn=lambda: scrape_coto(headless=headless),
        max_attempts=2,
    )
    print(f"  Total Coto: {len(prods)} productos")
    todos_los_productos.extend(prods)

    # La Anónima
    print("\n[La Anónima]")
    prods = _scrapear_cadena_con_recheck(
        cadena="La Anonima",
        tabla="precios",
        etiqueta="productos",
        scraper_fn=lambda: scrape_anonima(headless=headless),
    )
    print(f"  Total La Anónima: {len(prods)} productos")
    todos_los_productos.extend(prods)

    print(f"\nTotal general: {len(todos_los_productos)} productos")

    if not todos_los_productos:
        print("No se encontraron productos. Abortando.")
        return

    # Análisis de calidad + auto-corrección (antes de guardar)
    productos_corregidos = analizar_calidad(todos_los_productos)

    # Historial (con datos ya corregidos)
    historial = cargar_historial()
    agregar_corrida(historial, productos_corregidos)
    guardar_historial(historial)
    guardar_en_sqlite(productos_corregidos, str(date.today()))
    print(f"Historial guardado ({len(historial['semanas'])} corridas acumuladas)")

    # Excel
    generar_excel(historial)

    print("\nListo.")


if __name__ == "__main__":
    main()
